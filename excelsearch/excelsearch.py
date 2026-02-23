import ollama
import json
import math
import argparse
import os
import openpyxl

# --- Helper Functions ---
def dot_product(v1, v2):
    return sum(a * b for a, b in zip(v1, v2))

def magnitude(v):
    return math.sqrt(sum(a * a for a in v))

def cosine_similarity(v1, v2):
    return dot_product(v1, v2) / (magnitude(v1) * magnitude(v2))

# --- Convert a raw row dict into a natural language summary ---
def row_to_summary(raw):
    """
    Turns structured fields into a descriptive sentence the embedding
    model can reason about semantically.
    """
    parts = []

    vehicle = raw.get("Vehicle Type", "Vehicle")
    duration = raw.get("Duration (Mins)")
    fee = raw.get("Total Fee", "unknown fee")
    payment = raw.get("Payment Method", "unknown method")
    entry = raw.get("Entry Time", "")
    exit_ = raw.get("Exit Time", "")
    tid = raw.get("Transaction ID", "")

    # Duration in human terms
    if duration:
        try:
            mins = int(duration)
            hours = mins // 60
            remaining = mins % 60
            if hours > 0:
                duration_text = f"{hours} hour{'s' if hours > 1 else ''} and {remaining} minutes" if remaining else f"{hours} hour{'s' if hours > 1 else ''}"
            else:
                duration_text = f"{mins} minutes"

            # Tag it
            if mins >= 300:
                duration_text += " (very long stay)"
            elif mins >= 120:
                duration_text += " (long stay)"
            elif mins <= 30:
                duration_text += " (short stay)"
        except:
            duration_text = f"{duration} minutes"
    else:
        duration_text = "unknown duration"

    # Fee in human terms
    fee_text = fee
    if fee:
        try:
            amount = float(str(fee).replace("$", "").replace(",", ""))
            if amount == 0:
                fee_text = "free (validated or exempt)"
            elif amount >= 20:
                fee_text = f"${amount:.2f} (expensive)"
            elif amount >= 10:
                fee_text = f"${amount:.2f} (moderate cost)"
            else:
                fee_text = f"${amount:.2f} (cheap)"
        except:
            fee_text = fee

    parts.append(f"Transaction {tid}: A {vehicle} entered at {entry} and exited at {exit_}.")
    parts.append(f"The visit lasted {duration_text}.")
    parts.append(f"Total fee was {fee_text}, paid by {payment}.")

    return " ".join(parts)

# --- Load Excel File ---
def load_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else "" for h in rows[0]]
        for row in rows[1:]:
            if not any(row):
                continue
            raw = {headers[i]: str(val) for i, val in enumerate(row) if val is not None}
            original = " | ".join([f"{headers[i]}: {str(val)}" for i, val in enumerate(row) if val is not None])
            summary = row_to_summary(raw)   # <-- richer text for embedding
            chunks.append({
                "text": original,           # shown to user
                "summary": summary,         # used for embedding
                "raw": raw,
                "source": f"{filepath} [{sheet_name}]"
            })
    return chunks

# --- Embed and Save ---
def embed_excel(excel_files):
    data_to_store = []
    for filepath in excel_files:
        if not os.path.exists(filepath):
            print(f"Warning: {filepath} not found, skipping...")
            continue
        if not filepath.endswith(('.xlsx', '.xls')):
            print(f"Warning: {filepath} is not an Excel file, skipping...")
            continue
        print(f"Processing {filepath}...")
        chunks = load_excel(filepath)
        for chunk in chunks:
            print(f"  Embedding: {chunk['summary']}")
            response = ollama.embed(model='nomic-embed-text', input=chunk["summary"])
            vector = response['embeddings'][0]
            data_to_store.append({
                "text": chunk["text"],
                "summary": chunk["summary"],
                "raw": chunk["raw"],
                "source": chunk["source"],
                "vector": vector
            })
        print(f"  → {len(chunks)} rows embedded from {filepath}")

    with open("excel_vectors.json", "w") as f:
        json.dump(data_to_store, f)
    print(f"\nDone! Saved {len(data_to_store)} entries to excel_vectors.json")

# --- Exact match filter ---
def apply_filter(database, filter_str):
    filters = [f.strip() for f in filter_str.split(",")]
    filtered = []
    for entry in database:
        match = True
        for f in filters:
            if ":" not in f:
                continue
            field, value = f.split(":", 1)
            field, value = field.strip().lower(), value.strip().lower()
            raw = {k.lower(): v.lower() for k, v in entry["raw"].items()}
            if not any(field in k and value in v for k, v in raw.items()):
                match = False
                break
        if match:
            filtered.append(entry)
    return filtered

# --- Search ---
def search(query, top_k=3, filter_str=None, show_summary=False):
    if not os.path.exists("excel_vectors.json"):
        print("No excel_vectors.json found. Run with --index first.")
        return

    with open("excel_vectors.json", "r") as f:
        database = json.load(f)

    if filter_str:
        database = apply_filter(database, filter_str)
        print(f"Filter applied: {len(database)} rows matched")

    if not database:
        print("No results after filtering.")
        return

    response = ollama.embed(model='nomic-embed-text', input=query)
    query_vector = response['embeddings'][0]

    results = []
    for entry in database:
        score = cosine_similarity(query_vector, entry["vector"])
        results.append((score, entry["source"], entry["text"], entry.get("summary", "")))

    results.sort(reverse=True)
    print(f"\nSearch results for: '{query}'\n{'='*50}")
    for i, (score, source, text, summary) in enumerate(results[:top_k], 1):
        print(f"\nResult #{i} (score: {score:.4f})")
        print(f"Source: {source}")
        for field in text.split(" | "):
            print(f"  {field}")
        if show_summary:
            print(f"  [Embedded as]: {summary}")
        print("---")

# --- CLI ---
parser = argparse.ArgumentParser(description="Semantic search over Excel files")
parser.add_argument('--index', nargs='+', metavar='EXCEL_FILE',
                    help='Excel files to embed e.g. --index data.xlsx')
parser.add_argument('--query', type=str,
                    help='Search query e.g. --query "expensive trip"')
parser.add_argument('--filter', type=str,
                    help='Exact field filter e.g. --filter "Payment Method:Credit Card"')
parser.add_argument('--top', type=int, default=3,
                    help='Number of results to return (default: 3)')
parser.add_argument('--show-summary', action='store_true',
                    help='Show the natural language summary used for embedding')
args = parser.parse_args()

if args.index:
    embed_excel(args.index)
if args.query:
    search(args.query, top_k=args.top, filter_str=args.filter, show_summary=args.show_summary)
if not args.index and not args.query:
    parser.print_help()